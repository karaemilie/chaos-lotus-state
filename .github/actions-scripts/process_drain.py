"""
process_drain.py — runs inside GitHub Action

Reads drain.json (this repo), processes each completion against beast.xlsx
(in karaemilie/hiveBrain), writes beast back, clears drain.json.

After successful processing, calls the worker's /clear-uids endpoint to
remove processed items from KV (F9: prevents ghost re-injection).

Also processes drain.json's `pendingClear` field if present — that's how
Claude (chat-side) signals "please clear these uids from KV." Chat writes
the uids to that field; Action does the network call (Claude container
can't reach workers.dev).

MVP scope: ZONES + MAINTENANCE auto-process. SPIN_WHEEL/TASKS/COURAGE
completions stay queued for Claude.
"""

import os
import sys
import json
import time
import base64
import urllib.request
import urllib.error
import io
from datetime import datetime
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

# ─── CONFIG ──────────────────────────────────────────────────────
TOKEN = os.environ.get("CROSS_REPO_TOKEN", "")
if not TOKEN:
    print("❌ FATAL: CROSS_REPO_TOKEN env var not set")
    sys.exit(1)

BEAST_REPO = "karaemilie/hiveBrain"
BEAST_FILE = "masterHiveBrain.xlsx"
BEAST_BRANCH = "main"

DRAIN_PATH = "drain.json"  # in current repo (checkout)

WORKER_URL = "https://chaos-lotus-kv.theodidact.workers.dev"

# Sources this MVP handles automatically. All others stay queued for Claude.
# COURAGE stays manual: it routes a parent task to COMPLETED via the sacred
# process_completions pipeline, which does not live in the Action's world.
AUTO_SOURCES = {"ZONES", "MAINTENANCE", "SPIN_WHEEL"}


# ─── ALASKA STAMP ────────────────────────────────────────────────
def alaska_stamp_date():
    """Return today's date in Alaska time (with 3am pivot)."""
    now_ak = datetime.now(ZoneInfo("America/Anchorage"))
    if now_ak.hour < 3:
        from datetime import timedelta
        return (now_ak - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0, tzinfo=None)
    return now_ak.replace(hour=0, minute=0, second=0, microsecond=0, tzinfo=None)


# ─── GITHUB API ──────────────────────────────────────────────────
def gh(method, path, body=None):
    url = f"https://api.github.com{path}"
    headers = {
        "Authorization": f"Bearer {TOKEN}",
        "Accept": "application/vnd.github+json",
        "User-Agent": "process-drain-action",
    }
    data = json.dumps(body).encode() if body else None
    if data:
        headers["Content-Type"] = "application/json"
    req = urllib.request.Request(url, data=data, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            return resp.status, json.loads(resp.read().decode())
    except urllib.error.HTTPError as e:
        try:
            body = json.loads(e.read().decode())
        except Exception:
            body = {"error": "could not parse"}
        return e.code, body


def call_worker_clear_uids(uids):
    """Call worker POST /clear-uids to remove specific items from KV.

    Returns (status, response_dict). Best-effort — caller decides if a
    failure here is fatal (usually not — beast state is already correct,
    KV mismatch is recoverable).
    """
    if not uids:
        return 0, {"note": "no uids to clear"}
    body = json.dumps({"uids": list(uids)}).encode()
    req = urllib.request.Request(
        f"{WORKER_URL}/clear-uids",
        data=body,
        headers={
            "Content-Type": "application/json",
            "User-Agent": "process-drain-action",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=20) as resp:
            return resp.status, json.loads(resp.read().decode())
    except urllib.error.HTTPError as e:
        try:
            return e.code, json.loads(e.read().decode())
        except Exception:
            return e.code, {"error": "could not parse"}
    except Exception as e:
        return 0, {"error": str(e)}


# ─── LOAD/SAVE BEAST ─────────────────────────────────────────────
def load_beast():
    status, meta = gh("GET", f"/repos/{BEAST_REPO}/contents/{BEAST_FILE}?ref={BEAST_BRANCH}")
    if status != 200:
        raise RuntimeError(f"load_beast failed: HTTP {status} — {meta}")
    return base64.b64decode(meta["content"]), meta["sha"]


def save_beast(beast_bytes, sha, commit_msg):
    body = {
        "message": commit_msg,
        "content": base64.b64encode(beast_bytes).decode(),
        "branch": BEAST_BRANCH,
        "sha": sha,
    }
    status, result = gh("PUT", f"/repos/{BEAST_REPO}/contents/{BEAST_FILE}", body)
    if status not in (200, 201):
        raise RuntimeError(f"save_beast failed: HTTP {status} — {result}")
    return result


class BeastDeferred(Exception):
    """Raised when save_beast_with_retry exhausts its 409 budget under heavy
    concurrent load. This is NOT a real error: because the worker only clears
    drain.json AFTER successful processing, an exhausted run leaves its items
    IN the drain, so the next run reprocesses them (self-heal). main() catches
    this and exits 0 so it does NOT send a false-alarm 'All jobs failed' email
    for a condition that self-corrects. Genuine errors (auth, corruption,
    non-409) still raise RuntimeError and exit 1 — those need a human."""
    pass


def save_beast_with_retry(apply_fn, commit_msg, max_attempts=8):
    """Save the Beast with SHA-409 retry — the Beast analogue of the existing
    state.json retry. On 409, RELOAD the Beast fresh (new bytes + new sha),
    RE-APPLY all of this run's mutations via apply_fn (every processor keys by
    ID/ZID/SID/MID — never by row — so re-applying is idempotent), then retry.

    Under a heavy completion flood many runs fire concurrently and race for the
    Beast sha. If we exhaust the (now 8) attempts we raise BeastDeferred (NOT
    RuntimeError): the items stay queued in drain.json and the next run sweeps
    them, so no data is lost and no scary failure email fires. Jittered backoff
    spreads retries so concurrent runs don't keep colliding in lockstep.

    apply_fn signature: apply_fn(wb, stamp) -> None  (mutates wb in place)
    Returns (github commit result dict, final wb) on success.
    """
    import random
    attempt = 0
    while True:
        attempt += 1
        beast_bytes, beast_sha = load_beast()
        wb = load_workbook(io.BytesIO(beast_bytes))
        stamp = alaska_stamp_date()
        apply_fn(wb, stamp)
        buf = io.BytesIO(); wb.save(buf)
        body = {
            "message": commit_msg,
            "content": base64.b64encode(buf.getvalue()).decode(),
            "branch": BEAST_BRANCH,
            "sha": beast_sha,
        }
        status, result = gh("PUT", f"/repos/{BEAST_REPO}/contents/{BEAST_FILE}", body)
        if status in (200, 201):
            if attempt > 1:
                print(f"   ✅ Beast saved on attempt {attempt} (after 409 retry)")
            return result, wb
        if status == 409 and attempt < max_attempts:
            # jittered backoff: base grows with attempt, plus random spread so
            # concurrent runs stop colliding in lockstep.
            delay = 0.4 * attempt + random.uniform(0, 0.6)
            print(f"   ⚠️  Beast 409 conflict — reloading + re-applying ({attempt}/{max_attempts}, wait {delay:.1f}s)")
            time.sleep(delay)
            continue
        if status == 409:
            # Exhausted the budget under flood — defer, do NOT crash. Items stay
            # in drain (worker only clears after success) → next run reprocesses.
            print(f"   ⏸️  Beast still 409 after {attempt} attempts — DEFERRING this run's "
                  f"completions to the next run (items remain in drain, self-heal). No data lost.")
            raise BeastDeferred(f"409 after {attempt} attempts")
        # a non-409 error is a REAL problem self-heal won't fix → crash loud.
        raise RuntimeError(f"save_beast_with_retry failed: HTTP {status} after {attempt} attempt(s) — {result}")


# ─── ZONE PROCESSING ─────────────────────────────────────────────
def _uid_tail_int(uid, prefix):
    """Extract the trailing integer id from a uid, tolerant of how many
    colon-segments sit between the prefix and the id.

    Canonical forms are 2-part (ZONES:9, MAINTENANCE:4, SPIN:7), but an older
    front-end emitted 3-part ZONES:{floor}:{row} (e.g. 'ZONES:Personal:82').
    Taking the LAST segment as the id handles both without choking, so a
    format drift can never again silently jam the drain. Returns int or None.
    """
    if not uid or not uid.startswith(prefix + ":"):
        return None
    last = uid.split(":")[-1]
    try:
        return int(last)
    except (TypeError, ValueError):
        return None


def process_zone_completion(wb, comp, stamp):
    """Stamp a ZONES sheet cell. Returns (ok, message)."""
    floor = comp.get("floor")
    # ID-BASED: uid = ZONES:{zid}. Find the row whose ZID column == zid.
    # Tolerant parse: accepts canonical ZONES:{zid} AND legacy ZONES:{floor}:{row}
    # (in which case the trailing segment is the row, NOT a zid — so if a bare
    # 'zid' field isn't supplied we still try, then fall back to row-match).
    uid = comp.get("uid", "")
    zid = comp.get("zid")
    if zid is None:
        zid = _uid_tail_int(uid, "ZONES")
    if zid is None:
        return False, f"  ⚠️  malformed ZONE entry {uid!r}: no resolvable ZID"

    ws = wb["ZONES"]
    header = {c.value: i + 1 for i, c in enumerate(ws[1])}
    if "Completed" not in header or "ZID" not in header:
        return False, "  ❌ ZONES sheet missing Completed or ZID column"
    completed_col = header["Completed"]
    zid_col = header["ZID"]

    # Find the row by ZID (NOT by row number — rows shift, IDs don't)
    target_row = None
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, zid_col).value == zid:
            target_row = r
            break

    # LEGACY FALLBACK: a 3-part ZONES:{floor}:{row} uid's tail is a ROW, not a
    # ZID. If the ZID lookup missed AND the uid had 3 parts, try matching the
    # literal sheet row (and verify the floor name) so old drain entries clear.
    if target_row is None and uid.count(":") == 2:
        parts = uid.split(":")
        legacy_floor, legacy_row = parts[1], parts[2]
        try:
            lr = int(legacy_row)
            if 2 <= lr <= ws.max_row:
                row_floor = ws.cell(lr, header.get("Floor", 1)).value
                if row_floor == legacy_floor or floor == legacy_floor:
                    target_row = lr
        except (TypeError, ValueError):
            pass

    if target_row is None:
        return False, f"  ⚠️  ZID/row from {uid!r} not found in ZONES — SKIP"

    actual_floor = ws.cell(target_row, header["Floor"]).value
    actual_zone = ws.cell(target_row, header["Zone"]).value

    existing = ws.cell(target_row, completed_col).value
    if existing is not None:
        return True, f"  ⏭️  {actual_floor}/{actual_zone} (row {target_row}) already stamped — skip"

    ws.cell(target_row, completed_col).value = stamp
    return True, f"  ✅ {actual_floor}/{actual_zone} (row {target_row}) stamped {stamp.date()}"


def process_maintenance_completion(wb, comp, stamp):
    """Stamp a MAINTENANCE sheet row.

    MAINTENANCE schema: col 1=Order, col 2=Task, col 3=Completed.
    uid format: MAINTENANCE:{row}
    """
    # ID-BASED: uid = MAINTENANCE:{mid}. Find row whose MID column == mid.
    uid = comp.get("uid", "")
    mid = comp.get("mid")
    if mid is None:
        mid = _uid_tail_int(uid, "MAINTENANCE")
    if mid is None:
        return False, f"  ⚠️  malformed MAINTENANCE entry {uid!r}: no resolvable MID"

    ws = wb["MAINTENANCE"]
    header = {c.value: i + 1 for i, c in enumerate(ws[1])}
    if "Completed" not in header or "MID" not in header:
        return False, "  ❌ MAINTENANCE sheet missing Completed or MID column"
    completed_col = header["Completed"]
    mid_col = header["MID"]
    task_col = header.get("Task", 2)

    target_row = None
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, mid_col).value == mid:
            target_row = r
            break
    if target_row is None:
        return False, f"  ⚠️  MID {mid} not found in MAINTENANCE — SKIP"

    actual_task = ws.cell(target_row, task_col).value
    existing = ws.cell(target_row, completed_col).value
    if existing is not None:
        return True, f"  ⏭️  MAINTENANCE '{actual_task}' (MID {mid}) already stamped — skip"

    ws.cell(target_row, completed_col).value = stamp
    return True, f"  ✅ MAINTENANCE '{actual_task}' (MID {mid}) stamped {stamp.date()}"


def _log_spin_to_completed(wb, label, sid, stamp):
    """Append a finished spin task to the COMPLETED sheet so it shows up in
    'what did I do today'. Spin rows are minimal (Task/Source/TaskRow/SID), so
    we fill only the columns we have, by HEADER NAME (never positional):
      • ID            → f"SPIN-{sid}"  (namespaced: cannot collide with the
                        integer TASKS/COMPLETED ID space, so the ID audit stays clean)
      • Task          → the spin label
      • Category      → "Spin"   (so day-reports can group/identify spin wins)
      • Completed Date → Alaska stamp
      • Notes         → provenance breadcrumb
    Returns a log line, or None if COMPLETED sheet is absent.
    """
    if "COMPLETED" not in wb.sheetnames:
        return None
    wsc = wb["COMPLETED"]
    Hc = {c.value: i + 1 for i, c in enumerate(wsc[1])}

    # DEDUPE GUARD: if this SID is already logged with the SAME completion date,
    # don't append a second row. Without this, a completion that arrives via two
    # paths — e.g. a stranded retry-queue item that finally flushes PLUS a manual
    # re-mark of the same task — logs twice (the SPIN-27 / SPIN-150 double-logs).
    # Keyed on ID + Completed Date so the same task completed again on a LATER
    # day still logs (legitimately), but a same-day duplicate is skipped.
    _id = f"SPIN-{sid}"
    id_col = Hc.get("ID")
    date_col = Hc.get("Completed Date")
    if id_col and date_col:
        stamp_d = stamp.date() if hasattr(stamp, "date") else stamp
        for r in range(2, wsc.max_row + 1):
            if wsc.cell(r, id_col).value == _id:
                ev = wsc.cell(r, date_col).value
                ev_d = ev.date() if hasattr(ev, "date") else ev
                if ev_d == stamp_d:
                    return f"  ⏭️  SPIN-{sid} already logged {stamp_d} — skipping duplicate"

    nr = wsc.max_row + 1

    def put(col, val):
        if col in Hc:
            wsc.cell(nr, Hc[col]).value = val

    put("ID", f"SPIN-{sid}")
    put("Task", label)
    put("Category", "Spin")
    put("Completed Date", stamp)
    put("Notes", "✅ Completed via spin wheel")
    return f"  📝 logged to COMPLETED: '{label}' (SPIN-{sid})"


def process_spin_wheel_completions(wb, comps, stamp=None):
    """Log each completed spin task to COMPLETED, then delete its SPIN WHEEL row.

    uid format: SPIN_WHEEL:{row}. Deletes shift everything below up by one,
    so we MUST delete in descending row order or later deletes hit the wrong
    rows. Returns (processed_uids, messages).

    stamp: Alaska completion date (datetime). If None, completions are still
    deleted but NOT logged (caller should always pass it).
    """
    if "SPIN WHEEL" not in wb.sheetnames:
        return [], ["  ❌ SPIN WHEEL sheet not found — skipping all spin completions"]

    ws = wb["SPIN WHEEL"]
    header = {c.value: i + 1 for i, c in enumerate(ws[1])}
    if "SID" not in header:
        return [], ["  ❌ SPIN WHEEL missing SID column — cannot ID-match"]
    sid_col = header["SID"]
    task_col = header.get("Task", 1)
    processed = []
    msgs = []

    # ID-BASED: uid = SPIN:{sid}. Resolve each SID to its CURRENT row, then
    # delete by descending row so shifts don't corrupt remaining deletes.
    # We resolve rows fresh (not from drain) so prior deletes in THIS batch
    # are already reflected by re-scanning each time.
    targets = []  # (sid, uid, label)
    for comp in comps:
        uid = comp.get("uid", "")
        sid = comp.get("sid")
        if sid is None:
            # The front-end emits source 'SPIN_WHEEL' with uid 'SPIN_WHEEL:{sid}'.
            # Older notes/contract sometimes say 'SPIN:{sid}'. Accept BOTH prefixes
            # so a naming drift can't silently strand a spin completion.
            sid = _uid_tail_int(uid, "SPIN_WHEEL")
            if sid is None:
                sid = _uid_tail_int(uid, "SPIN")
        if sid is None:
            msgs.append(f"  ⚠️  malformed SPIN entry {uid!r}: no resolvable SID")
            continue
        targets.append((sid, uid, comp.get("label", "?")))

    # Resolve all SIDs to rows FIRST (before any deletes), then delete descending.
    sid_to_row = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, sid_col).value
        if v is not None:
            sid_to_row[v] = r

    resolved = []
    for sid, uid, label in targets:
        row = sid_to_row.get(sid)
        if row is None:
            # SID already absent from the sheet (e.g. an overlapping Action run
            # deleted the row first, or it shifted out). Previously this branch
            # cleared the uid as "done" but NEVER logged the win — so a spin
            # completion that resolved here vanished with no COMPLETED record
            # (the 'posykin handle' disappearance). We DON'T have the sheet row
            # anymore, but the drain payload carried the label, so log from that
            # before clearing. Without this, any racing/already-gone spin loses
            # its win silently.
            if stamp is not None and label and label != "?":
                logline = _log_spin_to_completed(wb, label, sid, stamp)
                if logline:
                    msgs.append(logline + " (from payload — row already gone)")
            msgs.append(f"  ⏭️  SID {sid} not found (already gone?) — logged from payload, clearing uid")
            processed.append(uid)  # already absent = effectively done
            continue
        resolved.append((row, sid, uid, label))

    for row, sid, uid, label in sorted(resolved, key=lambda x: x[0], reverse=True):
        actual = ws.cell(row, task_col).value
        # 1. log to COMPLETED first (so a delete failure can't lose the win)
        if stamp is not None:
            logline = _log_spin_to_completed(wb, actual or label, sid, stamp)
            if logline:
                msgs.append(logline)
        # 2. then delete the SPIN WHEEL row
        ws.delete_rows(row, 1)
        msgs.append(f"  ✅ SPIN deleted SID {sid} ('{actual or label}')")
        processed.append(uid)

    return processed, msgs


def process_adds(wb, adds):
    """Append each PLUS_ADD from the front end to the SPIN WHEEL sheet as a new
    one-off spin row. Per the 2026-06-14 design decision: everything added from
    a front-end app becomes a SPIN item — no classification, no Claude triage.

    Each new row gets a fresh SID (max existing + 1, incrementing across this
    batch). Returns (processed_add_keys, new_spin_items, messages) where:
      • processed_add_keys = the drain-add identity keys we handled (so they're
        dropped from drain.json)
      • new_spin_items = list of dicts for state.json/wheel refill seeding so the
        new item shows on the wheel immediately: {source,row(SID),label,uid,...}
      • messages = human-readable log lines
    """
    if not adds:
        return [], [], []
    if "SPIN WHEEL" not in wb.sheetnames:
        return [], [], ["  ❌ SPIN WHEEL sheet not found — cannot file adds"]

    ws = wb["SPIN WHEEL"]
    header = {c.value: i + 1 for i, c in enumerate(ws[1])}
    task_col = header.get("Task", 1)
    src_col = header.get("Source")
    sid_col = header.get("SID")
    if sid_col is None:
        return [], [], ["  ❌ SPIN WHEEL missing SID column — cannot file adds"]

    # Next SID = max existing + 1
    existing_sids = [ws.cell(r, sid_col).value for r in range(2, ws.max_row + 1)
                     if isinstance(ws.cell(r, sid_col).value, (int, float))]
    next_sid = (max(existing_sids) + 1) if existing_sids else 1

    processed_keys = []
    new_items = []
    msgs = []

    # De-dup guard: don't add a label that already sits on the SPIN WHEEL sheet
    existing_labels = {str(ws.cell(r, task_col).value).strip().lower()
                       for r in range(2, ws.max_row + 1)
                       if ws.cell(r, task_col).value}

    for add in adds:
        label = (add.get("label") or "").strip()
        added_at = add.get("_addedAt", "")
        # identity key matches how the front-end/worker keys a PLUS_ADD uid
        key = f"PLUS_ADD:{label}:{added_at}"
        if not label:
            msgs.append("  ⏭️  add with empty label — skipping")
            processed_keys.append(key)  # drop it from drain regardless
            continue
        if label.lower() in existing_labels:
            msgs.append(f"  ⏭️  '{label}' already on SPIN WHEEL — skip dup, clearing from drain")
            processed_keys.append(key)
            continue

        sid = next_sid
        next_sid += 1
        new_row = ws.max_row + 1
        ws.cell(new_row, task_col).value = label
        if src_col:
            ws.cell(new_row, src_col).value = "PLUS_ADD"
        ws.cell(new_row, sid_col).value = sid
        existing_labels.add(label.lower())

        # Wheel item so it shows immediately (uid uses canonical SPIN form)
        new_items.append({
            "source": "SPIN_WHEEL",
            "row": sid,           # SID is the stable key (front-end calls it row in spin uid)
            "sid": sid,
            "label": label,
            "uid": f"SPIN_WHEEL:{sid}",
        })
        processed_keys.append(key)
        msgs.append(f"  ✅ ADD filed → SPIN WHEEL SID {sid}: '{label}'")

    return processed_keys, new_items, msgs


def load_state_json():
    """Load state.json from THIS repo (chaos-lotus-state) via github API.
    Returns (state_dict, sha) or (None, None) if missing."""
    status, meta = gh("GET", "/repos/karaemilie/chaos-lotus-state/contents/state.json?ref=main")
    if status != 200:
        return None, None
    import base64
    return json.loads(base64.b64decode(meta["content"])), meta["sha"]


def save_state_json(state, sha, commit_msg):
    """Save state.json back to chaos-lotus-state repo, with SHA-conflict retry.

    Rapid completions trigger overlapping Action runs whose state.json writes
    race: the second writer's `sha` goes stale, GitHub returns 409 Conflict,
    and without a retry that write (INCLUDING its version bump) is silently
    lost — leaving the live state at a stale version, so front-ends that gate
    on `version` never see the update and keep showing completed items (ghosts).

    On conflict we re-fetch the current state, replay our task-set + a version
    bump ON TOP of whatever landed meanwhile, and retry — up to 4 times.
    """
    import base64

    def _put(body_state, body_sha):
        body = {
            "message": commit_msg,
            "content": base64.b64encode(json.dumps(body_state, indent=2, default=str).encode()).decode(),
            "branch": "main",
        }
        if body_sha:
            body["sha"] = body_sha
        return gh("PUT", "/repos/karaemilie/chaos-lotus-state/contents/state.json", body)

    status, result = _put(state, sha)
    attempts = 0
    while status == 409 and attempts < 4:
        attempts += 1
        print(f"   ⚠️  state.json 409 conflict — reloading + retrying ({attempts}/4)")
        fresh, fresh_sha = load_state_json()
        if fresh is None:
            break
        # Replay our intended task-set on top of fresh, bump version from fresh
        # so it always strictly increases past whatever just landed.
        merged = dict(fresh)
        merged["tasks"] = state.get("tasks", fresh.get("tasks", []))
        merged["version"] = fresh.get("version", 0) + 1
        merged["updated"] = state.get("updated")
        merged["buckets"] = state.get("buckets", fresh.get("buckets"))
        state = merged
        status, result = _put(state, fresh_sha)
    return status, result


def auto_reset_exhausted_floors(wb, floors):
    """Clear Completed dates for every zone on each exhausted floor, AND for
    MAINTENANCE if 'Maintenance' is flagged. Returns dict {floor: cleared_count}.

    This is the 'last mile' that makes floor-cycle resets fully hands-off: when
    refill_zone / refill_maintenance detect a wrapped (exhausted) floor, instead
    of only flagging resetNeeded for a human/Claude, the Action clears the cycle
    itself in the same run. Caller is responsible for committing the Beast after.
    """
    if not floors:
        return {}
    cleared = {}
    floors = set(floors)

    # ZONES floors
    if "ZONES" in wb.sheetnames:
        ws = wb["ZONES"]
        H = {c.value: i + 1 for i, c in enumerate(ws[1])}
        fcol, ccol = H.get("Floor"), H.get("Completed")
        if fcol and ccol:
            for r in range(2, ws.max_row + 1):
                fl = ws.cell(r, fcol).value
                if fl in floors and ws.cell(r, ccol).value is not None:
                    ws.cell(r, ccol).value = None
                    cleared[fl] = cleared.get(fl, 0) + 1

    # MAINTENANCE pseudo-floor
    if "Maintenance" in floors and "MAINTENANCE" in wb.sheetnames:
        ws = wb["MAINTENANCE"]
        H = {c.value: i + 1 for i, c in enumerate(ws[1])}
        ccol = H.get("Completed")
        if ccol:
            n = 0
            for r in range(2, ws.max_row + 1):
                if ws.cell(r, ccol).value is not None:
                    ws.cell(r, ccol).value = None
                    n += 1
            if n:
                cleared["Maintenance"] = n

    return cleared


def apply_refills(wb, processed_completions, today, pending_uids=None):
    """For each processed completion, remove it from state.json and add a fresh
    item in the same category. Returns (refill_summary, ok).

    processed_completions: list of the drain completion dicts that were
    successfully stamped/deleted this run.
    pending_uids: uids of completions still queued for Claude's engine (tapped
    but not yet finalized) — excluded from refill so they don't resurface.
    """
    pending_uids = pending_uids or set()
    try:
        import chaos_kv_refill as refill
    except ImportError:
        print("   ⚠️  chaos_kv_refill not importable — skipping refill")
        return [], False, []

    state, sha = load_state_json()
    if state is None:
        print("   ⚠️  state.json not loadable — skipping refill")
        return [], False, []

    tasks = state.get("tasks", [])
    completed_uids = {c.get("uid") for c in processed_completions}

    # IDEMPOTENCY GUARD (prevents over-fill from overlapping/queued Action runs):
    # When taps fire fast, multiple runs can each see the same completion in the
    # drain (the worker keeps appending; clears lag). Without this, every run
    # refills for every completion it sees → the wheel balloons (45 vs 42).
    # Fix: a completion only earns a refill if its item was ACTUALLY on the wheel
    # at the start of THIS run. If a prior run already removed it, it's already
    # been refilled — skip. We capture the pre-removal uid set to decide this.
    uids_before = {t.get("uid") for t in tasks}

    # 1. Remove completed items from the wheel
    tasks = [t for t in tasks if t.get("uid") not in completed_uids]

    # Only these completions get a refill: their item was present before removal.
    # (Spin/zone/maint matched by uid; tasks/courage matched by uid too.)
    refillable_uids = {u for u in completed_uids if u in uids_before}

    # 2. For each completion, compute a replacement (skip spin — by design).
    #    CRITICAL: exclude the just-completed uids AND their parent-task ids from
    #    eligibility, so a task you just finished can't immediately reappear.
    on_wheel = {t.get("uid") for t in tasks}
    # Build exclusion set covering all uid shapes a completed task could match:
    # its own uid, plus the cross-source forms (a completed COURAGE:123:0 must
    # also block TASKS:123, and vice-versa).
    just_done = set(completed_uids)
    for c in processed_completions:
        pid = c.get("parentId") or c.get("id")
        if pid is not None:
            just_done.add(f"TASKS:{pid}")
            just_done.add(f"COURAGE:{pid}:0")
    exclude = on_wheel | just_done | set(pending_uids)
    summary = []
    reset_floors = []  # floors whose Completed dates need a Claude-side reset
    for comp in processed_completions:
        # Idempotency: skip refill if this completion's item wasn't on the wheel
        # at the start of this run (a prior overlapping run already handled it).
        if comp.get("uid") not in refillable_uids:
            continue
        src = comp.get("source")
        # TRUE on-wheel TASKS count (recomputed each iteration as refills are
        # appended) — NOT the exclude set, which is polluted with just-completed
        # task uids. The cap-guard must measure the actual wheel so a batch of
        # completions can each refill instead of being blocked after the first.
        real_task_count = sum(1 for t in tasks if t.get("source") == "TASKS")
        new_item = refill.refill_for(src, wb, comp, exclude, today, wheel_task_count=real_task_count)
        if new_item:
            # Catch a floor-wrap reset signal (strip it before storing on wheel)
            rf_floor = new_item.pop("_resetFloor", None)
            if rf_floor:
                reset_floors.append(rf_floor)
            # DEDUP SPECIALS: never allow two Frogs or two Business items on the
            # wheel. Under rapid taps, overlapping runs could each add a special,
            # and an extra special inflates the TASKS count past the cap-guard's
            # headroom — silently blocking later daily-ten refills (the "41 short"
            # bug). If a special of this type is already present, skip the add.
            sp = new_item.get("specialZone")
            if sp and any(t.get("specialZone") == sp for t in tasks):
                summary.append(f"{src}: (skipped dup {sp})")
                continue
            tasks.append(new_item)
            exclude.add(new_item["uid"])
            pid2 = new_item.get("parentId") or new_item.get("id")
            if pid2 is not None:
                exclude.add(f"TASKS:{pid2}")
                exclude.add(f"COURAGE:{pid2}:0")
            summary.append(f"{src}: +{new_item['label'][:40]}")
        elif src != "SPIN_WHEEL":
            summary.append(f"{src}: (none eligible)")

    # 2b. SAFETY SWEEP: dedupe any specials that slipped in from a prior run
    # (keep the first of each specialZone). Belt-and-suspenders against the
    # cap-guard-inflation bug.
    seen_special = set()
    deduped = []
    for t in tasks:
        sp = t.get("specialZone")
        if sp:
            if sp in seen_special:
                continue  # drop the duplicate special
            seen_special.add(sp)
        deduped.append(t)
    tasks = deduped

    # 3. Write state.json back with bumped version
    state["tasks"] = tasks
    state["version"] = state.get("version", 0) + 1
    state["updated"] = datetime.now(ZoneInfo("UTC")).isoformat()
    counts = {}
    for t in tasks:
        counts[t.get("source", "?")] = counts.get(t.get("source", "?"), 0) + 1
    state["buckets"] = counts

    status, result = save_state_json(
        state, sha,
        f"🔄 Auto-refill: -{len(completed_uids)} done, +{len([s for s in summary if s.startswith(('ZONES','MAINTENANCE','TASKS','COURAGE')) and '+' in s])} fresh"
    )
    if status in (200, 201):
        print(f"   ✅ state.json refilled → version {state['version']}")
        return summary, True, reset_floors
    else:
        print(f"   ⚠️  state.json save failed: HTTP {status}")
        return summary, False, reset_floors


def finalize_task_completions(comps):
    """Finalize TASKS + COURAGE completions in the Beast via the canonical
    process_completions engine (single source of truth — co-located in this
    repo, so the Action always runs the current version).

    COURAGE:{parentId}:0 and TASKS:{id} both resolve to a TASKS row by ID;
    completing a courage item completes its parent task. Returns
    (results_dict_or_None, ok).
    """
    if not comps:
        return None, True
    try:
        import process_completions as pcmod
    except ImportError:
        print("   ⚠️  process_completions engine not importable — leaving TASKS/COURAGE queued")
        return None, False

    # Resolve each uid to a parent TASKS id
    ids = []
    for c in comps:
        uid = c.get("uid", "")
        if c.get("source") == "TASKS":
            tid = c.get("id")
            if tid is None and uid.startswith("TASKS:"):
                try: tid = int(uid.split(":")[1])
                except (IndexError, ValueError): tid = None
            if tid is not None: ids.append(tid)
        elif c.get("source") == "COURAGE":
            pid = c.get("parentId")
            if pid is None and uid.startswith("COURAGE:"):
                try: pid = int(uid.split(":")[1])
                except (IndexError, ValueError): pid = None
            if pid is not None: ids.append(pid)
    if not ids:
        return None, True

    print(f"\n📜 Finalizing {len(ids)} TASKS/COURAGE completion(s) via engine: {ids}")
    beast_bytes, sha = load_beast()
    wb = load_workbook(io.BytesIO(beast_bytes))
    results = pcmod.process_completions(wb, ids)

    for tid, name in results.get("completed", []):
        print(f"   ✅ {name[:45]} → COMPLETED")
    for nid, name, nxt in results.get("recurred", []):
        print(f"   🔄 recurring next: {name[:35]} → {nxt}")
    for uid_, name in results.get("unlocked", []):
        print(f"   🔓 unlocked next-in-seq: {name[:40]}")
    if results.get("errors"):
        print(f"   ⚠️  errors: {results['errors']}")
    if results.get("audit"):
        print(f"   🔴 AUDIT PROBLEMS: {results['audit']} — NOT saving, leaving queued")
        return results, False  # do not save a beast that failed audit

    # Save beast back
    buf = io.BytesIO(); wb.save(buf)
    save_beast(buf.getvalue(), sha, f"🤖 Auto-finalize: {len(results.get('completed',[]))} TASKS/COURAGE → COMPLETED")
    print("   🐝 Beast updated with finalized completions")
    return results, True


def write_sync_receipt(processed_uids, processed_details, remaining_completions, remaining_adds):
    """Write a token-free verification breadcrumb to the PUBLIC state repo.

    Chat-side Claude can read this via raw.githubusercontent.com with no token,
    so it can confirm 'did my tap file correctly?' without ever touching the
    private Beast. The Beast stays private; only this summary is public.

    Committed by the workflow yaml alongside drain.json.
    """
    receipt = {
        "lastRun": datetime.now(ZoneInfo("UTC")).isoformat(),
        "lastRunAlaska": datetime.now(ZoneInfo("America/Anchorage")).strftime("%Y-%m-%d %H:%M %Z"),
        "processedCount": len(processed_uids),
        "processed": processed_details,   # list of {uid, result} human-readable
        "remainingForClaude": {
            "completions": len(remaining_completions),
            "adds": len(remaining_adds),
        },
    }
    with open("last_sync.json", "w") as f:
        json.dump(receipt, f, indent=2)
    print(f"\n🧾 Sync receipt written: {len(processed_uids)} processed, "
          f"{len(remaining_completions)} completions + {len(remaining_adds)} adds left for Claude")


# ─── MAIN ────────────────────────────────────────────────────────
def main():
    # 1. Read drain.json from local checkout
    try:
        with open(DRAIN_PATH, "r") as f:
            drain = json.load(f)
    except FileNotFoundError:
        print(f"⚠️  {DRAIN_PATH} not found — nothing to process")
        return

    completions = drain.get("completions", [])
    adds = drain.get("adds", [])
    pending_clear_from_chat = drain.get("pendingClear", [])
    print(f"📥 Drain has {len(completions)} completions + {len(adds)} adds + {len(pending_clear_from_chat)} pendingClear uids")

    if not completions and not adds and not pending_clear_from_chat:
        print("✨ Drain is empty — exiting cleanly")
        return

    # 2. Partition by auto-handleable vs manual
    auto_zones = [c for c in completions if c.get("source") == "ZONES"]
    auto_maintenance = [c for c in completions if c.get("source") == "MAINTENANCE"]

    # A spin completion is anything whose SOURCE is SPIN_WHEEL *or* whose UID is a
    # spin uid (SPIN:{sid} / SPIN_WHEEL:{sid}). Routing on uid-prefix as a fallback
    # closes the PLUS_ADD stranding bug: a +added petal that reached the wheel
    # carrying its sheet provenance ("PLUS_ADD") as source would otherwise fail the
    # source=="SPIN_WHEEL" gate, fall into other_completions, never delete its row,
    # and resurface forever on every refill (the "condition leather" loop).
    def _is_spin(c):
        if c.get("source") == "SPIN_WHEEL":
            return True
        uid = c.get("uid", "") or ""
        return uid.startswith("SPIN:") or uid.startswith("SPIN_WHEEL:")

    auto_spin = [c for c in completions if _is_spin(c)]
    # Everything that is neither an auto source NOR a spin-uid completion is left
    # for Claude. (Spin items are excluded even when their source tag is wrong.)
    other_completions = [c for c in completions
                         if c.get("source") not in AUTO_SOURCES and not _is_spin(c)]

    print(f"🔧 Auto-processable: {len(auto_zones)} ZONES, {len(auto_maintenance)} MAINTENANCE, {len(auto_spin)} SPIN_WHEEL, {len(adds)} ADDS→SPIN")
    print(f"⏸️  Leaving for Claude: {len(other_completions)} other completions")

    if pending_clear_from_chat:
        print(f"📞 Chat-side pendingClear: {len(pending_clear_from_chat)} uid(s) to clear from KV")

    # 3. Load beast if we have ANY beast-touching work: auto completions OR adds
    #    (adds get appended to the SPIN WHEEL sheet).
    wb = None
    beast_sha = None
    if auto_zones or auto_maintenance or auto_spin or adds:
        print(f"\n📂 Loading beast from {BEAST_REPO}/{BEAST_FILE}...")
        beast_bytes, beast_sha = load_beast()
        print(f"   {len(beast_bytes)} bytes, SHA {beast_sha[:12]}")
        wb = load_workbook(io.BytesIO(beast_bytes))
    stamp = alaska_stamp_date() if wb else None
    if wb:
        print(f"   Stamp date (Alaska): {stamp.date()}")

    # 4a. Process ZONE completions
    processed_uids = []
    processed_details = []  # human-readable lines for the public receipt
    processed_comps = []    # the actual completion dicts that succeeded (for refill)
    if auto_zones and wb:
        print(f"\n🏠 Processing {len(auto_zones)} ZONE completions:")
        for comp in auto_zones:
            ok, msg = process_zone_completion(wb, comp, stamp)
            print(msg)
            if ok:
                processed_uids.append(comp.get("uid"))
                processed_details.append({"uid": comp.get("uid"), "result": msg.strip()})
                processed_comps.append(comp)

    # 4b. Process MAINTENANCE completions
    if auto_maintenance and wb:
        print(f"\n🔧 Processing {len(auto_maintenance)} MAINTENANCE completions:")
        for comp in auto_maintenance:
            ok, msg = process_maintenance_completion(wb, comp, stamp)
            print(msg)
            if ok:
                processed_uids.append(comp.get("uid"))
                processed_details.append({"uid": comp.get("uid"), "result": msg.strip()})
                processed_comps.append(comp)

    # 4c. Process SPIN_WHEEL completions (ID-based SID delete)
    if auto_spin and wb:
        print(f"\n🎡 Processing {len(auto_spin)} SPIN_WHEEL completions:")
        spin_uids, spin_msgs = process_spin_wheel_completions(wb, auto_spin, stamp=stamp)
        for m in spin_msgs:
            print(m)
        processed_uids.extend(spin_uids)
        for u, m in zip(spin_uids, [x for x in spin_msgs if x.strip().startswith("✅")]):
            processed_details.append({"uid": u, "result": m.strip()})
        # track spin comps for refill bookkeeping (refill itself skips spin)
        for comp in auto_spin:
            if comp.get("uid") in spin_uids:
                processed_comps.append(comp)

    # 4d. Process ADDS → append to SPIN WHEEL sheet (front-end adds become spin items)
    new_spin_items = []
    processed_add_keys = []
    if adds and wb:
        print(f"\n➕ Processing {len(adds)} ADD(s) → SPIN WHEEL:")
        processed_add_keys, new_spin_items, add_msgs = process_adds(wb, adds)
        for m in add_msgs:
            print(m)

    beast_dirty = bool(processed_uids) or bool(processed_add_keys)
    if not beast_dirty:
        print("\n⏭️  No auto-completions or adds applied this run")
    else:
        # 5. Save beast — via SHA-409 retry. The replay closure re-applies EVERY
        # mutation this run made (zones/maint/spin completions + adds) onto a
        # FRESH Beast if a concurrent run bumped the sha. All processors key by
        # ID/ZID/SID/MID (never row), so replaying onto a changed Beast is safe
        # + idempotent. This closes the drop where a 409 used to crash the run
        # after the drain moved on, losing a completion's stamp entirely.
        print(f"\n💾 Saving beast back to github (with 409-retry)...")
        commit_msg = f"🤖 Auto-process: {len(processed_uids)} completion(s) + {len(processed_add_keys)} add(s)"

        def _replay_mutations(_wb, _stamp):
            # Re-apply zone completions
            for _c in auto_zones:
                if _c.get("uid") in processed_uids:
                    process_zone_completion(_wb, _c, _stamp)
            # Re-apply maintenance completions
            for _c in auto_maintenance:
                if _c.get("uid") in processed_uids:
                    process_maintenance_completion(_wb, _c, _stamp)
            # Re-apply spin deletions AND logging. MUST pass _stamp — without it,
            # _log_spin_to_completed is gated off and the replayed (= actually
            # saved) Beast loses the COMPLETED record even though the first in-
            # memory pass logged it. This was the silent spin-win disappearance
            # (posykin, ultravine): fixed in the first copy, dropped in the replay.
            _spin_to_apply = [_c for _c in auto_spin if _c.get("uid") in processed_uids]
            if _spin_to_apply:
                process_spin_wheel_completions(_wb, _spin_to_apply, stamp=_stamp)
            # Re-apply adds (process_adds dedupes by label, so replaying is safe)
            if adds:
                process_adds(_wb, adds)

        result, wb = save_beast_with_retry(_replay_mutations, commit_msg)
        print(f"   Committed: {result['commit']['sha'][:12]}")

    # Uids of everything still queued for Claude (TASKS/COURAGE not finalized) —
    # exclude these from refill so a tapped-but-unprocessed task can't resurface.
    all_reset_floors = []  # floors that wrapped and need a Claude-side reset
    pending_queue_uids = {c.get("uid") for c in other_completions}
    # also block their parent-id cross-forms
    for c in other_completions:
        pid = c.get("parentId") or c.get("id")
        if pid is not None:
            pending_queue_uids.add(f"TASKS:{pid}")
            pending_queue_uids.add(f"COURAGE:{pid}:0")

    # 5b+5c MERGED: refill ZONES/MAINT/SPIN *and* TASKS/COURAGE in ONE state
    # read-modify-write. Previously these were two separate apply_refills calls,
    # each independently loading + saving state.json. In a batch run (mixed zone
    # + task completions) they raced: the 2nd call loaded state BEFORE the 1st
    # saved, so its write clobbered the 1st's removals/refills — leaving a
    # completed zone stuck on the wheel (zombie) AND dropping refills (wheel
    # short of target). One combined call = one atomic state cycle = no race.
    tc_comps = [c for c in other_completions if c.get("source") in ("TASKS", "COURAGE")]
    refill_only = [c for c in tc_comps if not c.get("_refilled")]

    # Ensure we have a workbook for the TASKS/COURAGE refill eligibility lookups.
    wb_refill = wb
    if wb_refill is None and refill_only:
        try:
            rb_bytes, _ = load_beast()
            wb_refill = load_workbook(io.BytesIO(rb_bytes))
        except Exception as e:
            print(f"   ⚠️  couldn't load beast for refill: {e}")
            wb_refill = None

    # Combined completion set for a single refill pass. processed_comps =
    # ZONES/MAINT/SPIN that auto-processed; refill_only = TASKS/COURAGE.
    all_refill_comps = list(processed_comps) + list(refill_only)
    if all_refill_comps and (wb_refill is not None or wb is not None):
        print(f"\n🔄 Refilling wheel ({len(processed_comps)} zone/maint/spin + {len(refill_only)} task/courage)...")
        today_ak = alaska_stamp_date().date()
        # pending = TASKS/COURAGE not yet finalized this run (none here, since
        # we refill them now) — keep empty so their refills aren't excluded.
        combined_summary, refill_ok, combined_resets = apply_refills(
            wb_refill if wb_refill is not None else wb,
            all_refill_comps, today_ak, pending_uids=set(),
        )
        all_reset_floors += combined_resets
        for line in combined_summary:
            print(f"   {line}")

    # 5a-RESET: AUTO-RESET exhausted floors (the 'last mile' — fully hands-off).
    # refill_zone/refill_maintenance set _resetFloor when a floor wraps (every
    # zone done). Rather than only flagging resetNeeded for a human, clear the
    # floor's Completed dates HERE, commit the Beast, and re-seed that floor's
    # fresh zones onto the wheel so they appear immediately.
    auto_resolved_floors = []
    if all_reset_floors and (wb_refill is not None or wb is not None):
        reset_wb = wb_refill if wb_refill is not None else wb
        cleared = auto_reset_exhausted_floors(reset_wb, set(all_reset_floors))
        if cleared:
            print(f"\n♻️  AUTO-RESET exhausted floor(s): {cleared}")
            # Second Beast commit carrying the cleared cycle.
            try:
                buf = io.BytesIO(); reset_wb.save(buf)
                # re-read current sha to avoid conflict with the earlier save
                _b, cur_sha = load_beast()
                rmsg = "♻️ Auto-reset exhausted floor(s): " + ", ".join(
                    f"{k}(-{v})" for k, v in cleared.items())
                rres = save_beast(buf.getvalue(), cur_sha, rmsg)
                print(f"   Committed reset: {rres['commit']['sha'][:12]}")
                auto_resolved_floors = list(cleared.keys())
                # Re-seed the freshly-reset zones onto the wheel right away.
                try:
                    import chaos_kv_refill as _rf
                    state2, sha2 = load_state_json()
                    if state2 is not None:
                        on_wheel2 = {t.get("uid") for t in state2.get("tasks", [])}
                        added2 = 0
                        for fl in cleared:
                            if fl == "Maintenance":
                                continue
                            seeded = _rf.refill_zone(
                                reset_wb, {"uid": "", "source": "ZONES", "floor": fl},
                                on_wheel2)
                            if seeded:
                                seeded.pop("_resetFloor", None)
                                if seeded["uid"] not in on_wheel2:
                                    state2["tasks"].append(seeded)
                                    on_wheel2.add(seeded["uid"])
                                    added2 += 1
                        if added2:
                            state2["version"] = state2.get("version", 0) + 1
                            state2["updated"] = datetime.now(ZoneInfo("UTC")).isoformat()
                            counts2 = {}
                            for t in state2["tasks"]:
                                counts2[t.get("source", "?")] = counts2.get(t.get("source", "?"), 0) + 1
                            state2["buckets"] = counts2
                            save_state_json(state2, sha2,
                                            f"♻️ Seed {added2} reset-floor zone(s) onto wheel")
                            print(f"   🪷 Seeded {added2} fresh zone(s) post-reset")
                except Exception as e:
                    print(f"   ⚠️  post-reset reseed skipped: {e}")
            except Exception as e:
                print(f"   ⚠️  auto-reset Beast commit failed ({e}); leaving resetNeeded flag")
                auto_resolved_floors = []

    # Now finalize ALL pending TASKS/COURAGE in the Beast (including any from
    # prior runs that were refilled but not yet finalized).
    fin_results, fin_ok = finalize_task_completions(tc_comps)
    if fin_ok and tc_comps:
        # mark these uids as fully processed so they're cleared + removed from drain
        for c in tc_comps:
            u = c.get("uid")
            if u and u not in processed_uids:
                processed_uids.append(u)

    # 6. Determine what to clear from KV:
    #    - everything we just processed (completions)
    #    - every ADD we filed to the SPIN WHEEL sheet (so KV stops re-mirroring
    #      it — otherwise a filed add lingers as a ghost and can re-fire)
    #    - everything chat-side flagged via drain.json's `pendingClear` field
    all_clear_uids = list(set(processed_uids) | set(processed_add_keys) | set(pending_clear_from_chat))

    if all_clear_uids:
        print(f"\n🧹 Calling worker /clear-uids for {len(all_clear_uids)} uid(s):")
        for u in all_clear_uids:
            print(f"   - {u}")
        clear_status, clear_result = call_worker_clear_uids(all_clear_uids)
        print(f"   Status: {clear_status}")
        if clear_status == 200:
            print(f"   ✅ Removed: {clear_result.get('removed')} | Remaining: {clear_result.get('remaining')}")
        else:
            print(f"   ⚠️  Worker /clear-uids non-200: {clear_result}")
            print(f"   Beast state is still correct; KV may have ghosts until next run.")

    # 7. Rewrite drain.json — remove processed items, drop pendingClear field.
    #    TASKS/COURAGE that we refilled this run get a `_refilled` flag so the
    #    next run won't refill them again (they stay queued for Claude's engine).
    refilled_uids = {c.get("uid") for c in refill_only}
    new_completions = []
    for c in completions:
        if c.get("uid") in processed_uids:
            continue  # fully done — drop
        if c.get("uid") in refilled_uids:
            c = {**c, "_refilled": True}  # mark so we don't double-refill
        new_completions.append(c)

    # ADDS: drop the ones we filed into the SPIN WHEEL sheet (matched by identity
    # key PLUS_ADD:{label}:{_addedAt}). Anything not processed stays queued.
    processed_add_set = set(processed_add_keys)
    new_adds = []
    for a in adds:
        key = f"PLUS_ADD:{(a.get('label') or '').strip()}:{a.get('_addedAt','')}"
        if key in processed_add_set:
            continue  # filed → drop
        new_adds.append(a)

    new_drain = {
        **drain,
        "completions": new_completions,
        "adds": new_adds,
        "processedAt": datetime.now(ZoneInfo("UTC")).isoformat(),
    }
    if processed_uids:
        new_drain["lastAutoProcessed"] = processed_uids
    # Floors that wrapped this run. Any we AUTO-RESET above are recorded as
    # resolved; only floors we couldn't auto-clear stay flagged for fallback.
    still_need = [f for f in set(all_reset_floors) if f not in set(auto_resolved_floors)]
    if still_need:
        existing = set(new_drain.get("resetNeeded", []))
        new_drain["resetNeeded"] = sorted(existing | set(still_need))
        print(f"   🔄 Flagged floors for Claude-side reset: {new_drain['resetNeeded']}")
    if auto_resolved_floors:
        new_drain["autoReset"] = sorted(set(auto_resolved_floors))
        # clear any stale resetNeeded entries we just handled
        if "resetNeeded" in new_drain:
            new_drain["resetNeeded"] = sorted(
                set(new_drain["resetNeeded"]) - set(auto_resolved_floors))
            if not new_drain["resetNeeded"]:
                new_drain.pop("resetNeeded", None)
        print(f"   ♻️  Auto-reset floors this run: {new_drain['autoReset']}")
    # Remove pendingClear field once we've acted on it
    new_drain.pop("pendingClear", None)

    with open(DRAIN_PATH, "w") as f:
        json.dump(new_drain, f, indent=2)
    print(f"\n🧹 Drain rewritten: {len(new_completions)} completions remaining, {len(new_adds)} adds remaining")
    print(f"   (workflow yaml will commit drain.json change)")

    # 7b. Seed newly-filed spin adds into state.json so they appear on the wheel
    #     immediately (without waiting for a full re-seed). Best-effort.
    if new_spin_items:
        try:
            state, sha = load_state_json()
            if state is not None:
                on_wheel = {t.get("uid") for t in state.get("tasks", [])}
                added = 0
                for item in new_spin_items:
                    if item["uid"] not in on_wheel:
                        state.setdefault("tasks", []).append(item)
                        added += 1
                if added:
                    state["version"] = state.get("version", 0) + 1
                    state["updated"] = datetime.now(ZoneInfo("UTC")).isoformat()
                    counts = {}
                    for t in state["tasks"]:
                        counts[t.get("source", "?")] = counts.get(t.get("source", "?"), 0) + 1
                    state["buckets"] = counts
                    st, _ = save_state_json(state, sha, f"➕ Seed {added} new spin add(s) onto wheel")
                    print(f"   🪷 Seeded {added} new spin add(s) onto wheel (state save HTTP {st})")
        except Exception as e:
            print(f"   ⚠️  could not seed new spin adds into state.json: {e}")

    # 8. Write public sync receipt (token-free verification for chat-side Claude)
    write_sync_receipt(processed_uids, processed_details, new_completions, new_adds)


if __name__ == "__main__":
    try:
        main()
    except BeastDeferred as e:
        # Expected under heavy concurrent load — items stay in drain, next run
        # reprocesses them. Exit 0 so GitHub does NOT email "All jobs failed"
        # for a self-correcting condition.
        print(f"\n⏸️  Run deferred (no data lost, will self-heal next run): {e}")
        sys.exit(0)
