"""
═══════════════════════════════════════════════════════════════
  ZONE_PICKER.py — Deterministic Zone Assignment Engine
  Save this file to the Claude Project as a project file.
  Claude MUST run this script before assigning any zones.
  DO NOT FREESTYLE ZONES. RUN THIS. USE ITS OUTPUT.
═══════════════════════════════════════════════════════════════

HOW THIS WORKS:
  1. Opens masterHiveBrain.xlsx (ZONES sheet + MAINTENANCE sheet)
  2. Walks each floor's rows TOP TO BOTTOM
  3. Stops at the FIRST row with an empty 'Completed' cell
  4. That's the zone. No sorting. No filtering. No thinking.
  5. Applies catch-up rule for Digital, Plant, Personal
  6. Includes Maintenance zone from MAINTENANCE sheet
  7. Includes Business zone placeholder (Claude picks from TASKS)
  8. Prints clean output Claude can copy into reports

CATCH-UP RULE (Digital, Personal, Maintenance):
  If the floor's most recent Completed date ≠ yesterday,
  show TWO zones (first AND second empty rows) to prevent stacking.

ALWAYS-TWO RULE (Plant):
  Plant ALWAYS shows TWO zones regardless of when last done.
  If only 1 empty row remains, partial auto-reset: take the last
  empty row + reset the floor and grab row 1 of the new cycle.

AUTO-RESET RULE:
  If a floor has NO empty Completed cells, clear ALL values for that
  floor (dates AND X marks) and restart from the top. X means "skipped
  this cycle only" — NOT permanent. Clear with cell.value = None, then
  SAVE + RELOAD to verify. Never ask — just do it.

USAGE:
  Claude copies this script to /home/claude/, updates the xlsx path
  and today's date, runs it, and uses the output verbatim.

CRITICAL RULES:
  - FIRST EMPTY ROW = TODAY'S ZONE. Period.
  - Do NOT sort by date. Do NOT prioritize "never done."
  - The rotation order is built into the ROW ORDER of the sheet.
  - Walk down. Stop at blank. That's it.
  - Maintenance = first empty row in MAINTENANCE sheet.
  - Business = Claude picks ONE available Business task from TASKS
    (tagged 🏆 Q2). This is the only zone that requires judgment.
"""

import openpyxl
from datetime import datetime, timedelta
import sys
sys.path.insert(0, '/mnt/project')
from FLOOR_EMOJI import FLOOR_EMOJI as _FLOOR_EMOJI_SOURCE

def _execute_floor_reset(xlsx_path, floor_name, rows, is_done_today_fn, picked_rows=None):
    """2026-05-11 PATCH: actually clear the xlsx when a floor needs auto-reset.
    
    Opens a writable workbook, clears the Completed cell for every row on
    this floor EXCEPT:
      • rows marked done today (preserve today's stamp)
      • rows in picked_rows (these are the new picks — leave empty so they
        can be marked when completed)
    
    Note: rows already empty stay empty. Cleared cells were previously holding
    stale completion dates or X skips from the prior cycle.
    
    Side-effects: writes + saves the xlsx. Caller's read-only wb is unaffected.
    """
    import openpyxl
    if picked_rows is None:
        picked_rows = set()
    wb_write = openpyxl.load_workbook(xlsx_path)
    ws_write = wb_write['ZONES']
    cleared = 0
    for row_num, _zone, comp in rows:
        # Preserve today's stamp — it's a recent completion from this cycle
        if is_done_today_fn(comp):
            continue
        # Picked rows should be empty (None) so they can be marked when done
        # — if they were already None this is a no-op, but we want them empty
        # so they appear "fresh" in any reload
        cur = ws_write.cell(row_num, 3).value
        if cur is not None:
            ws_write.cell(row_num, 3).value = None
            cleared += 1
    if cleared > 0:
        wb_write.save(xlsx_path)
        print(f"  🔄 Auto-reset cleared {cleared} cells on {floor_name} floor")
    wb_write.close()


def pick_zones(xlsx_path, today=None, exclude_rows=None, force_full_sweep=False):
    """
    Deterministic zone picker. Returns dict of zone assignments.
    
    Args:
        xlsx_path: Path to masterHiveBrain.xlsx
        today: datetime object for today (defaults to now)
        exclude_rows: set of TASKS row numbers to skip when picking
                      Business / Frog zones. Prevents a task from appearing
                      on both the task side (spin wheel) AND zone side of
                      the same build. Pass the set of rows already staged
                      in SPIN WHEEL sheet.
        force_full_sweep: when True, bypass the "subtract done_today from quota"
                      logic. Every floor gets its full daily quota of fresh
                      zones regardless of what was already stamped today.
                      Use for SPIN WHEEL builds (the wheel IS the tool for
                      doing zones — don't pre-deduct progress).
                      Default False = standard subtract-done-today behavior
                      (use for daily reports that should reflect real status).
    
    Returns:
        dict with zone assignments per floor
    """
    if exclude_rows is None:
        exclude_rows = set()
    else:
        exclude_rows = set(exclude_rows)
    if today is None:
        # 3AM DAY BOUNDARY: Anything completed between midnight and 3am
        # counts toward the PREVIOUS day's quota. Night owl friendly.
        now = datetime.now()
        if now.hour < 3:
            today = (now - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
        else:
            today = now.replace(hour=0, minute=0, second=0, microsecond=0)
    
    yesterday = today - timedelta(days=1)
    
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb['ZONES']
    
    # Also read TASKS sheet for Business zone
    ws_tasks = wb['TASKS']
    headers = [ws_tasks.cell(1, c).value for c in range(1, ws_tasks.max_column + 1)]
    
    # ═══════════════════════════════════════
    # STEP 1: Read all zone rows per floor, IN ORDER
    # ═══════════════════════════════════════
    floors = {}  # {floor_name: [(row_num, zone_name, completed_date), ...]}
    
    for r in range(2, ws.max_row + 1):
        floor = ws.cell(r, 1).value
        zone = ws.cell(r, 2).value
        completed = ws.cell(r, 3).value
        
        if not floor or not zone:
            continue
        
        # Skip the Maintenance placeholder row
        if floor == 'Maintenance':
            continue
            
        if floor not in floors:
            floors[floor] = []
        floors[floor].append((r, zone, completed))
    
    # ═══════════════════════════════════════
    # STEP 2: Catch-up floors (show TWO zones if last done ≠ yesterday)
    #         ALWAYS_TWO floors show TWO zones unconditionally.
    # ═══════════════════════════════════════
    CATCHUP_FLOORS = {'Digital', 'Personal'}
    ALWAYS_TWO_FLOORS = {'Plant'}  # Plant ALWAYS shows 2 — no catch-up check
    
    # ═══════════════════════════════════════
    # STEP 3: For each floor, find first empty row(s)
    # ═══════════════════════════════════════
    assignments = {}
    
    for floor_name, rows in floors.items():
        # ═══════════════════════════════════════
        # "DONE TODAY" CHECK — the fix for phantom-zone bug.
        # Count how many zones on this floor were completed TODAY.
        # 3AM RULE: A completion counts as "today" only if its timestamp is
        # >= today_3am AND < tomorrow_3am. Date-only completions (hour=0)
        # fall back to date comparison since we can't tell when in the day
        # they happened.
        today_3am = today.replace(hour=3)
        tomorrow_3am = today_3am + timedelta(days=1)
        def _is_done_today(comp):
            if comp is None or not isinstance(comp, datetime):
                return False
            # Has real hour info → use 3am window
            if comp.hour != 0 or comp.minute != 0:
                return today_3am <= comp < tomorrow_3am
            # Date-only → fall back to calendar date match
            return comp.date() == today.date()
        done_today_count = sum(1 for _, _, comp in rows if _is_done_today(comp))
        
        # Find all empty rows (None = empty, 'X' = skipped this cycle only)
        # X marks are NOT permanent — they get cleared on auto-reset
        empty_rows = [(r, zone) for r, zone, comp in rows if comp is None]
        
        # AUTO-RESET: If no empty rows, clear ALL (dates AND X marks) and restart
        # X = "skipped this cycle only" — NOT permanent. Must be cleared on reset.
        # BUT: if the daily quota is already met by today's completions, show nothing.
        if not empty_rows:
            # Figure out daily quota for this floor so we can check if done_today
            # already satisfies it (auto-reset shouldn't fire if so).
            always_two_here = floor_name in ALWAYS_TWO_FLOORS
            daily_quota = 2 if always_two_here else 1
            
            if done_today_count >= daily_quota:
                # Floor is already satisfied for today — don't reset yet.
                assignments[floor_name] = {
                    'zones': [],
                    'auto_reset': False,
                    'partial_reset': False,
                    'catchup': False,
                    'always_two': always_two_here,
                    'last_done': None,
                    'done_today': done_today_count,
                    'satisfied': True,
                }
                continue
            
            # NOTE: When executing the auto-reset in the xlsx, clear ALL completed
            # cells for this floor including 'X' values. Use cell.value = None
            # explicitly, then SAVE and RELOAD to verify the clear took effect.
            #
            # 2026-05-11 PATCH (round 2): pick the rows with the OLDEST completion
            # dates, not rows[0]. Reasoning: when a full cycle completes, the
            # first rows in the floor list are usually the ones JUST done — so
            # picking them again on day-1 of new cycle = bug.
            #
            # X-skip semantics: X means "deliberately skipped this cycle".
            # Treat as settled-but-not-done. After auto-reset clears the cell,
            # the row is eligible again — but on the SAME reset call we
            # shouldn't immediately pick X rows. They sort LAST so we pick
            # actual dated rows (oldest-first) before falling back to X rows.
            remaining_after_done = max(0, daily_quota - done_today_count)
            
            def _sort_key(row_tuple):
                """Oldest dated first, then None, then X. done_today excluded by caller."""
                _r, _z, comp = row_tuple
                if isinstance(comp, datetime):
                    return (0, comp.timestamp())   # dated → sort by date (oldest first)
                if comp is None:
                    return (1, 0)                   # None → after dated
                if isinstance(comp, str):           # 'X' → deliberate skip, picks last
                    return (2, 0)
                return (3, 0)                       # unknown type → very last
            
            reset_candidates = [rt for rt in rows if not _is_done_today(rt[2])]
            reset_candidates.sort(key=_sort_key)
            picks = reset_candidates[:remaining_after_done]
            reset_zones = [(rt[0], rt[1]) for rt in picks]
            
            assignments[floor_name] = {
                'zones': reset_zones,
                'auto_reset': True,
                'always_two': always_two_here,
                'partial_reset': False,
                'catchup': False,
                'last_done': None,
                'done_today': done_today_count,
            }
            # Clear the xlsx: wipe all completed cells on this floor EXCEPT
            # done-today (preserve today's stamp) AND the rows we just picked
            # (which will get their new stamp when user completes them).
            picked_rows = {rt[0] for rt in picks}
            _execute_floor_reset(xlsx_path, floor_name, rows, _is_done_today, picked_rows)
            continue
        
        # Check if this floor needs catch-up or ALWAYS gets two
        needs_catchup = False
        always_two = floor_name in ALWAYS_TWO_FLOORS
        last_completed_date = None
        
        if floor_name in CATCHUP_FLOORS:
            # Find the most recent completed date for this floor
            completed_dates = [comp for _, _, comp in rows 
                             if comp is not None and isinstance(comp, datetime)]
            if completed_dates:
                last_completed_date = max(completed_dates)
                # Caught up if last done is yesterday OR today.
                # (Today counts — finishing today's zone means you're not behind.)
                lcd = last_completed_date.date()
                if lcd != yesterday.date() and lcd != today.date():
                    needs_catchup = True
            else:
                # Never done at all = definitely needs catch-up
                needs_catchup = True
        
        # Pick zone(s): ALWAYS_TWO floors get 2 unconditionally,
        # catch-up floors get 2 if triggered, others get 1
        num_to_pick = 2 if (always_two or needs_catchup) else 1
        
        # SUBTRACT what's already done today from the daily quota.
        # If done_today >= num_to_pick, this floor is SATISFIED for today.
        # EXCEPTION: force_full_sweep skips this — give the full quota
        # regardless of progress (used by wheel builds).
        if force_full_sweep:
            remaining_to_pick = num_to_pick
        else:
            remaining_to_pick = max(0, num_to_pick - done_today_count)
        
        if remaining_to_pick == 0:
            # Floor is done for today — show empty zones list.
            # Still record state flags so downstream knows it was satisfied.
            assignments[floor_name] = {
                'zones': [],
                'auto_reset': False,
                'partial_reset': False,
                'catchup': needs_catchup,
                'always_two': always_two,
                'last_done': last_completed_date,
                'done_today': done_today_count,
                'satisfied': True,
            }
            continue
        
        picked = empty_rows[:remaining_to_pick]
        
        # PARTIAL AUTO-RESET: If we need more than we have empty, take
        # what's left, then reset the floor and grab the first row of
        # the new cycle to fill the gap.
        partial_reset = False
        if remaining_to_pick >= 2 and len(picked) < remaining_to_pick:
            partial_reset = True
            # The first row in the floor's rotation becomes the extra pick
            first_row = rows[0]
            picked.append((first_row[0], first_row[1]))
        
        assignments[floor_name] = {
            'zones': picked,
            'auto_reset': False,
            'partial_reset': partial_reset,
            'catchup': needs_catchup,
            'always_two': always_two,
            'last_done': last_completed_date,
            'done_today': done_today_count,
        }
    
    # ═══════════════════════════════════════
    # STEP 4: Maintenance zone (from MAINTENANCE sheet)
    # Same first-empty-row logic + catch-up rule + done-today check
    # ═══════════════════════════════════════
    maintenance_picks = []
    maintenance_last_done = None
    maintenance_needs_catchup = False
    maintenance_done_today = 0
    
    try:
        ws_m = wb['MAINTENANCE']
        
        # Find the most recent completed date in Maintenance
        # AND count how many were done TODAY
        for r in range(2, ws_m.max_row + 1):
            completed = ws_m.cell(r, 3).value
            if completed and isinstance(completed, datetime):
                if maintenance_last_done is None or completed > maintenance_last_done:
                    maintenance_last_done = completed
                if completed.date() == today.date():
                    maintenance_done_today += 1
        
        # Check catch-up: if last done ≠ yesterday AND ≠ today, show TWO
        # (today counts — finishing a maintenance task today means you're not behind)
        if maintenance_last_done:
            lcd = maintenance_last_done.date()
            if lcd != yesterday.date() and lcd != today.date():
                maintenance_needs_catchup = True
        else:
            maintenance_needs_catchup = True  # Never done = catch up
        
        num_to_pick = 2  # Maintenance daily quota = 2
        
        # Subtract what's already done today from the quota
        # EXCEPTION: force_full_sweep skips subtraction (wheel build path)
        if force_full_sweep:
            remaining_to_pick = num_to_pick
        else:
            remaining_to_pick = max(0, num_to_pick - maintenance_done_today)
        
        # Walk rows, find first empty(s) — only if we still need to pick
        if remaining_to_pick > 0:
            for r in range(2, ws_m.max_row + 1):
                task = ws_m.cell(r, 2).value
                completed = ws_m.cell(r, 3).value
                if task and completed is None:
                    maintenance_picks.append((r, task))
                    if len(maintenance_picks) >= remaining_to_pick:
                        break
            
            if not maintenance_picks:
                # All done — auto-reset: pick first row
                maintenance_picks = [(2, ws_m.cell(2, 2).value)]
                assignments['_maintenance_reset'] = True
    except Exception:
        pass
    
    assignments['Maintenance'] = {
        'zones': maintenance_picks,
        'auto_reset': len(maintenance_picks) == 0 and maintenance_done_today < 2,
        'catchup': maintenance_needs_catchup and len(maintenance_picks) > 1,
        'last_done': maintenance_last_done,
        'done_today': maintenance_done_today,
        'satisfied': maintenance_done_today >= 2,
    }
    
    # ═══════════════════════════════════════
    # STEP 5: Business zone (from TASKS sheet)
    # Pull ONE available Business task. If none active today,
    # pull the earliest future Business task anyway.
    # Business = always Q2 🏆, always visible.
    # ═══════════════════════════════════════
    business_pick = None
    business_is_future = False
    
    try:
        active_biz = []
        future_biz = []
        
        for r in range(2, ws_tasks.max_row + 1):
            # DEDUP: skip tasks already on the spin wheel so we don't double-inject
            if r in exclude_rows:
                continue
            row = {}
            for c in range(1, ws_tasks.max_column + 1):
                row[headers[c-1]] = ws_tasks.cell(r, c).value
            
            if row.get('Category') != 'Business':
                continue
            
            start = row.get('Start Date')
            if not start or not isinstance(start, datetime):
                continue
            
            task_info = (r, row['Task'], row.get('Anchor Weight'), start)
            
            if start.date() <= today.date():
                active_biz.append(task_info)
            else:
                future_biz.append(task_info)
        
        if active_biz:
            # Pick first active Business task
            business_pick = active_biz[0]
            business_is_future = False
        elif future_biz:
            # No active — pull earliest future
            future_biz.sort(key=lambda x: x[3])
            business_pick = future_biz[0]
            business_is_future = True
    except Exception:
        pass
    
    assignments['Business'] = {
        'zones': [(business_pick[0], business_pick[1])] if business_pick else [],
        'auto_reset': False,
        'catchup': False,
        'last_done': None,
        'is_future': business_is_future,
        'start_date': business_pick[3] if business_pick else None,
        'aw': business_pick[2] if business_pick else None,
    }
    
    # ═══════════════════════════════════════
    # STEP 6: Frog zone (from TASKS sheet)
    # Smart tiered selection:
    #   Tier 1: Most overdue AW1/AW2 task (Due < today, Start ≤ today)
    #   Tier 2: Oldest Start Date task (any AW, Start ≤ today)
    #   Tier 3: Nothing eligible → graceful skip
    # 
    # Exclusions (never froggable):
    #   - Already on spin wheel (exclude_rows)
    #   - Already picked for Business zone (prevent double-dip)
    #   - Category == 'Business' (has its own zone)
    #   - Recurring tasks (frogs are one-off confrontations)
    #   - Sequential tasks not at order 1 (blocked tasks aren't frogs)
    # ═══════════════════════════════════════
    frog_excludes = set(exclude_rows)
    if business_pick:
        frog_excludes.add(business_pick[0])
    
    frog_pick = None           # tuple (row, task, aw, start, due, days_overdue)
    frog_tier = None           # 'overdue_priority' | 'oldest_start'
    
    try:
        tier1_candidates = []   # overdue AW1/AW2
        tier2_candidates = []   # any AW, Start ≤ today
        
        for r in range(2, ws_tasks.max_row + 1):
            if r in frog_excludes:
                continue
            row = {}
            for c in range(1, ws_tasks.max_column + 1):
                row[headers[c-1]] = ws_tasks.cell(r, c).value
            
            name = row.get('Task')
            if not name or not str(name).strip():
                continue
            
            # Exclude Business category (has its own zone)
            if row.get('Category') == 'Business':
                continue
            
            # Exclude actual recurring tasks (frogs are one-off confrontations).
            # Reads the migrated 'Recurring Type' field (Type='None' = not recurring).
            # Falls back to legacy 'Recurring' if Type absent (transitional safety).
            rec_type = row.get('Recurring Type')
            if rec_type is not None:
                if str(rec_type).strip().lower() not in ('none', '', 'one-time', 'false', 'no'):
                    continue
            else:
                recurring = row.get('Recurring')
                if recurring:
                    rec_str = str(recurring).strip().lower()
                    if rec_str and rec_str not in ('one-time', 'none', 'false', 'no', 'n/a'):
                        continue
            
            # Exclude blocked sequential tasks (only order 1 or null is froggable)
            seq_order = row.get('Sequential Order')
            if seq_order is not None:
                try:
                    if int(seq_order) > 1:
                        continue
                except (TypeError, ValueError):
                    pass
            
            start = row.get('Start Date')
            if not start or not isinstance(start, datetime):
                continue
            if start.date() > today.date():
                continue  # future task — not surfaceable
            
            aw = row.get('Anchor Weight')
            try:
                aw_int = int(aw) if aw is not None else None
            except (TypeError, ValueError):
                aw_int = None
            
            due = row.get('Due Date')
            if due and isinstance(due, datetime) and due.date() < today.date():
                days_overdue = (today.date() - due.date()).days
                if aw_int in (1, 2):
                    tier1_candidates.append((r, name, aw_int, start, due, days_overdue))
            
            # Every start-eligible task is a tier2 fallback candidate
            tier2_candidates.append((r, name, aw_int, start, due if isinstance(due, datetime) else None, 0))
        
        if tier1_candidates:
            # Sort by days_overdue DESC (most overdue first), tiebreak by AW
            tier1_candidates.sort(key=lambda x: (-x[5], x[2] or 99))
            frog_pick = tier1_candidates[0]
            frog_tier = 'overdue_priority'
        elif tier2_candidates:
            # Sort by Start Date ASC (oldest first), tiebreak by AW
            tier2_candidates.sort(key=lambda x: (x[3], x[2] or 99))
            frog_pick = tier2_candidates[0]
            frog_tier = 'oldest_start'
    except Exception:
        pass
    
    assignments['Frog'] = {
        'zones': [(frog_pick[0], frog_pick[1])] if frog_pick else [],
        'auto_reset': False,
        'catchup': False,
        'last_done': None,
        'tier': frog_tier,
        'aw': frog_pick[2] if frog_pick else None,
        'start_date': frog_pick[3] if frog_pick else None,
        'due_date': frog_pick[4] if frog_pick else None,
        'days_overdue': frog_pick[5] if frog_pick else None,
    }
    
    wb.close()
    return assignments


def format_output(assignments, today=None):
    """Pretty-print zone assignments for Claude to use in reports."""
    if today is None:
        today = datetime.now()
    
    FLOOR_EMOJI = _FLOOR_EMOJI_SOURCE  # single source of truth: /mnt/project/FLOOR_EMOJI.py
    
    # Display order
    DISPLAY_ORDER = ['Upstairs', 'Main Floor', 'Basement', 'Digital', 
                     'Plant', 'Personal', 'Maintenance', 'Business', 'Frog']
    
    print("=" * 60)
    print(f"  🏠 ZONE ASSIGNMENTS — {today.strftime('%A %B %d, %Y')}")
    print("=" * 60)
    
    total_zones = 0
    
    for floor in DISPLAY_ORDER:
        emoji = FLOOR_EMOJI.get(floor, '📍')
        
        if floor == 'Business':
            if floor in assignments and assignments[floor]['zones']:
                data = assignments[floor]
                zone_name = data['zones'][0][1]
                aw = data.get('aw', '?')
                if data.get('is_future'):
                    start = data.get('start_date')
                    start_str = start.strftime('%m/%d') if start else '?'
                    print(f"\n  {emoji} Business: 🏆 AW{aw} | {zone_name} (starts {start_str} — pulled early!)")
                else:
                    print(f"\n  {emoji} Business: 🏆 AW{aw} | {zone_name}")
            else:
                print(f"\n  {emoji} Business: 🏆 No Business tasks found!")
            total_zones += 1
            continue
        
        if floor == 'Frog':
            if floor in assignments and assignments[floor]['zones']:
                data = assignments[floor]
                zone_name = data['zones'][0][1]
                aw = data.get('aw', '?')
                tier = data.get('tier')
                if tier == 'overdue_priority':
                    days = data.get('days_overdue', '?')
                    print(f"\n  {emoji} Frog: 🐸 AW{aw} | {zone_name} (overdue -{days}d — EAT IT)")
                else:
                    start = data.get('start_date')
                    start_str = start.strftime('%m/%d') if start else '?'
                    print(f"\n  {emoji} Frog: 🐸 AW{aw} | {zone_name} (oldest start {start_str})")
            else:
                print(f"\n  {emoji} Frog: 🐸 No eligible frogs today — lucky you!")
            total_zones += 1
            continue
        
        if floor not in assignments:
            print(f"\n  {emoji} {floor}: ⚠️ NOT FOUND IN SHEET")
            continue
        
        data = assignments[floor]
        zones = data['zones']
        
        if data['auto_reset']:
            if data.get('always_two') and len(zones) > 1:
                print(f"\n  {emoji} {floor} (🔄 AUTO-RESET × 2 — new cycle!):")
                for i, (row, zone) in enumerate(zones):
                    label = "→ Zone" if i == 0 else "→ Zone 2"
                    print(f"      {label}: {zone} (row {row})")
                total_zones += 2
            else:
                print(f"\n  {emoji} {floor}: 🔄 AUTO-RESET (all done!) → {zones[0][1] if zones else 'N/A'}")
                total_zones += 1
            continue
        
        if data['catchup'] and len(zones) > 1:
            print(f"\n  {emoji} {floor} (CATCH-UP × 2 — last done: {data['last_done'].strftime('%m/%d') if data['last_done'] else 'never'}):")
            for i, (row, zone) in enumerate(zones):
                label = "→ Zone" if i == 0 else "→ Catch-up"
                print(f"      {label}: {zone} (row {row})")
            total_zones += 2
        elif data.get('always_two') and len(zones) > 1:
            reset_note = " 🔄 NEW CYCLE" if data.get('partial_reset') else ""
            print(f"\n  {emoji} {floor} (ALWAYS × 2):")
            for i, (row, zone) in enumerate(zones):
                label = "→ Zone" if i == 0 else f"→ Zone 2{reset_note}"
                print(f"      {label}: {zone} (row {row})")
            total_zones += 2
        else:
            zone_name = zones[0][1] if zones else 'N/A'
            row_num = zones[0][0] if zones else '?'
            
            if floor == 'Maintenance':
                if len(zones) > 1:
                    print(f"\n  {emoji} Maintenance:")
                    for i, (row, zone) in enumerate(zones):
                        label = "→ Zone" if i == 0 else "→ Zone 2"
                        print(f"      {label}: {zone} (row {row})")
                else:
                    print(f"\n  {emoji} Maintenance: {zone_name}")
            else:
                print(f"\n  {emoji} {floor}: {zone_name} (row {row_num})")
            total_zones += 1
    
    print(f"\n{'=' * 60}")
    print(f"  TOTAL ZONES: {total_zones}")
    print(f"  (Catch-up floors show 2 when last done ≠ yesterday)")
    print(f"{'=' * 60}")
    
    return total_zones


# ═══════════════════════════════════════
# RUN IT
# ═══════════════════════════════════════
if __name__ == "__main__":
    # Claude updates these before running:
    XLSX_PATH = "/mnt/project/masterHiveBrain.xlsx"
    TODAY = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    
    assignments = pick_zones(XLSX_PATH, TODAY)
    total = format_output(assignments, TODAY)
