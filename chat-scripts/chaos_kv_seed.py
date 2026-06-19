"""
chaos_kv_seed.py — Canonical wheel seed builder

Builds state.json for the chaos lotus wheel from the current beast.
Replaces ad-hoc inline seed code with a reusable, schema-aware helper.

CRITICAL: this is the SEND side. The companion file process_drain.py (in
.github/actions-scripts/) is the RECEIVE side. Both must agree on
source/uid conventions:

  source           uid format                stamp location
  ─────────────────────────────────────────────────────────────────────
  TASKS            TASKS:{id}                COMPLETED sheet (via process_completions)
  ZONES            ZONES:{floor}:{row}       ZONES sheet, Completed col
  MAINTENANCE      MAINTENANCE:{row}         MAINTENANCE sheet, Completed col
  SPIN_WHEEL       SPIN_WHEEL:{row}          SPIN WHEEL sheet row (delete on done)
  COURAGE          COURAGE:{parentId}:0      via parent TASKS row + process_completions
  PLUS_ADD         PLUS_ADD:{label}          New row in SPIN WHEEL or TASKS (TBD by Claude)
  PLUS_ADD_TASK    PLUS_ADD_TASK:{label}    New TASKS row, Claude classifies

Usage:
  from chaos_kv_seed import build_seed, push_seed, seed_wheel

  # All-in-one:
  seed_wheel()  # uses default beast path, today's date

  # Or step-by-step for inspection:
  tasks = build_seed()
  print(f"Built {len(tasks)} items")
  push_seed(tasks)
"""

import sys
import random
from datetime import date, datetime, timezone

# Ensure project files are importable (FLOOR_EMOJI, ZONE_PICKER, etc.)
sys.path.insert(0, "/mnt/project")

from openpyxl import load_workbook
from FLOOR_EMOJI import FLOOR_EMOJI
from ZONE_PICKER import pick_zones


# ─── CONFIG ──────────────────────────────────────────────────────
DEFAULT_BEAST_PATH = "/home/claude/masterHiveBrain.xlsx"

DAILY_TEN_COUNT = 15
COURAGE_COUNT = 3
DAILY_AW_RANGE = (3, 7)
COURAGE_AW_RANGE = (1, 2)


# ─── HELPERS ─────────────────────────────────────────────────────
def _coerce_date(s):
    """Parse a string Start/Due date into a date; return None if unparseable
    (so a stray text date skips rather than crashing the whole seed). Handles
    the common ISO and US formats that show up from various add paths."""
    if s is None:
        return None
    s = str(s).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%m/%d/%Y", "%m/%d/%y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _load_available_tasks(beast_path, today):
    """Return all TASKS rows where Start Date <= today. Each as a dict."""
    wb = load_workbook(beast_path, data_only=True)
    ws = wb["TASKS"]
    H = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    available = []
    for r in range(2, ws.max_row + 1):
        task = ws.cell(r, H["Task"]).value
        if not task:
            continue
        start = ws.cell(r, H["Start Date"]).value
        if hasattr(start, "date"):
            start = start.date()
        elif isinstance(start, str):
            # A Start Date stored as text (some add paths write strings) would
            # crash the date comparison below. Coerce common formats; if it
            # won't parse, treat as no-start (skip) rather than crash the seed.
            start = _coerce_date(start)
        if start is None or start > today:
            continue
        aw = ws.cell(r, H["Anchor Weight"]).value
        try:
            aw_int = int(aw) if aw is not None else None
        except Exception:
            aw_int = None
        if aw_int is None:
            continue

        due = ws.cell(r, H["Due Date"]).value
        if hasattr(due, "date"):
            due = due.date()

        rec_type = ws.cell(r, H["Recurring Type"]).value if "Recurring Type" in H else None
        is_recurring = bool(rec_type and str(rec_type).strip().lower() not in ("none", ""))

        available.append({
            "id": ws.cell(r, H["ID"]).value,
            "row": r,
            "label": task,
            "aw": aw_int,
            "pri": ws.cell(r, H["Priority"]).value,
            "dur": ws.cell(r, H["Duration (min)"]).value,
            "ml": ws.cell(r, H["Mental Load"]).value,
            "proj": ws.cell(r, H["Project"]).value,
            "cat": ws.cell(r, H["Category"]).value,
            "notes": ws.cell(r, H["Notes"]).value,
            "start": start,
            "due": due,
            "critical": bool(ws.cell(r, H["Critical"]).value),
            "recurring": is_recurring,
        })
    return available


def _load_future_onetime(beast_path, today, exclude_rows=None):
    """Pull-forward pool: one-time TASKS with Start Date STRICTLY AFTER today,
    nearest-upcoming first. RECURRING TASKS ARE EXCLUDED — they must never
    surface before their own Start date (completing one early would wrongly
    advance its cycle). Used only to top up Daily Ten / Courage when today's
    eligible pool is under cap. Same row-dict shape as _load_available_tasks.
    """
    exclude_rows = exclude_rows or set()
    wb = load_workbook(beast_path, data_only=True)
    ws = wb["TASKS"]
    H = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    future = []
    for r in range(2, ws.max_row + 1):
        if r in exclude_rows:
            continue
        task = ws.cell(r, H["Task"]).value
        if not task:
            continue
        start = ws.cell(r, H["Start Date"]).value
        if hasattr(start, "date"):
            start = start.date()
        elif isinstance(start, str):
            start = _coerce_date(start)
        if start is None or start <= today:   # only FUTURE-dated
            continue
        rec_type = ws.cell(r, H["Recurring Type"]).value if "Recurring Type" in H else None
        is_recurring = bool(rec_type and str(rec_type).strip().lower() not in ("none", ""))
        if is_recurring:                       # NEVER pull recurring forward
            continue
        aw = ws.cell(r, H["Anchor Weight"]).value
        try:
            aw_int = int(aw) if aw is not None else None
        except Exception:
            aw_int = None
        if aw_int is None:
            continue
        due = ws.cell(r, H["Due Date"]).value
        if hasattr(due, "date"):
            due = due.date()
        future.append({
            "id": ws.cell(r, H["ID"]).value,
            "row": r,
            "label": task,
            "aw": aw_int,
            "pri": ws.cell(r, H["Priority"]).value,
            "dur": ws.cell(r, H["Duration (min)"]).value,
            "ml": ws.cell(r, H["Mental Load"]).value,
            "proj": ws.cell(r, H["Project"]).value,
            "cat": ws.cell(r, H["Category"]).value,
            "notes": ws.cell(r, H["Notes"]).value,
            "start": start,
            "due": due,
            "critical": bool(ws.cell(r, H["Critical"]).value),
            "recurring": False,
            "pulled_forward": True,            # tag so callers/UI can tell
        })
    # nearest-upcoming first (tomorrow before next week), AW as light tiebreak
    future.sort(key=lambda t: (t["start"], -t["aw"]))
    return future


def _build_daily_ten(available, exclude_rows, today, beast_path=None):
    """Daily Ten: surface ALL due-today recurring (AW3-7) first — they're the
    easy, must-keep-up rhythm wins — then random-fill one-time tasks up to the
    cap. If recurring alone exceeds the cap, show them all anyway (rhythm tasks
    are never hidden). Daily-stable random seed for the fill. If STILL under cap
    after today's pool, pull FUTURE one-time tasks forward (nearest dates first,
    never recurring) so a light today doesn't leave the wheel half-empty."""
    lo, hi = DAILY_AW_RANGE
    pool = [t for t in available if lo <= t["aw"] <= hi and t["row"] not in exclude_rows]
    recurring = [t for t in pool if t.get("recurring")]
    onetime = [t for t in pool if not t.get("recurring")]
    picks = list(recurring)  # always surface every due recurring task
    slots_left = DAILY_TEN_COUNT - len(picks)
    if slots_left > 0 and onetime:
        rng = random.Random(today.toordinal())
        picks.extend(rng.sample(onetime, min(slots_left, len(onetime))))
    # PULL-FORWARD: still under cap? Top up from future one-time tasks,
    # nearest-upcoming first. Never recurring (handled by the loader).
    slots_left = DAILY_TEN_COUNT - len(picks)
    if slots_left > 0 and beast_path is not None:
        claimed = set(exclude_rows) | {t["row"] for t in picks}
        future = [t for t in _load_future_onetime(beast_path, today, claimed)
                  if lo <= t["aw"] <= hi]
        picks.extend(future[:slots_left])
    return picks


def _build_courage(available, exclude_rows, today, courage_micros=None, beast_path=None):
    """Most-overdue 3 from AW 1-2. If under cap, pull FUTURE one-time AW1-2 tasks
    forward (nearest dates first, never recurring) rather than the old weak
    oldest-start-from-today fallback."""
    courage_micros = courage_micros or {}
    lo, hi = COURAGE_AW_RANGE
    pool = [t for t in available if lo <= t["aw"] <= hi and t["row"] not in exclude_rows]
    overdue = [t for t in pool if t["due"] and t["due"] < today]
    overdue.sort(key=lambda t: (today - t["due"]).days, reverse=True)
    picks = list(overdue[:COURAGE_COUNT])
    if len(picks) < COURAGE_COUNT:
        # first exhaust any remaining due-today AW1-2 (oldest start first)
        rem = [t for t in pool if t not in picks]
        rem.sort(key=lambda t: t["start"])
        picks.extend(rem[:COURAGE_COUNT - len(picks)])
    if len(picks) < COURAGE_COUNT and beast_path is not None:
        # still short → pull future one-time AW1-2 forward, nearest first
        claimed = set(exclude_rows) | {t["row"] for t in picks}
        future = [t for t in _load_future_onetime(beast_path, today, claimed)
                  if lo <= t["aw"] <= hi]
        picks.extend(future[:COURAGE_COUNT - len(picks)])
    return picks


def _build_zones(beast_path, today):
    """Run ZONE_PICKER and split returns by FLOOR.

    Returns (zone_items, claimed_rows_for_tasks).

    Important nuance:
      • Returns for 'Upstairs/Main Floor/Basement/Digital/Plant/Personal' come
        from the ZONES sheet — these get source='ZONES', uid='ZONES:{floor}:{row}'.
      • Returns for 'Maintenance' come from the MAINTENANCE sheet (different sheet!) —
        these get source='MAINTENANCE', uid='MAINTENANCE:{row}'.
      • Returns for 'Business' and 'Frog' are TASKS sheet rows — these get
        source='TASKS', uid='TASKS:{id}' (so we don't duplicate them in Daily Ten).
        We track those rows in claimed_rows for downstream exclusion.
    """
    today_dt = datetime.combine(today, datetime.min.time())
    assignments = pick_zones(beast_path, today_dt, force_full_sweep=True)

    # Build row→ID lookup maps from the ZID/MID columns (appended, schema-safe).
    _wb_ids = load_workbook(beast_path, data_only=True)
    def _row_id_map(sheet, id_header):
        ws = _wb_ids[sheet]
        H = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        col = H.get(id_header)
        return {r: ws.cell(r, col).value for r in range(2, ws.max_row + 1)} if col else {}
    ZID_BY_ROW = _row_id_map("ZONES", "ZID")
    MID_BY_ROW = _row_id_map("MAINTENANCE", "MID")

    items = []
    claimed_rows = set()

    # ZONES sheet floors
    for floor in ('Upstairs', 'Main Floor', 'Basement', 'Digital', 'Plant', 'Personal'):
        if floor not in assignments:
            continue
        emoji = FLOOR_EMOJI.get(floor, '📍')
        for row, zone_name in assignments[floor].get('zones', []):
            zid = ZID_BY_ROW.get(row)
            if zid is None:
                continue  # no ZID — skip rather than emit unsafe row uid
            items.append({
                "source": "ZONES",
                "zid": zid,
                "floor": floor,
                "row": row,            # reference only — NOT the key
                "label": f"{emoji} {zone_name}",
                "emoji": emoji,
                "zoneName": zone_name,
                "uid": f"ZONES:{zid}",
            })

    # MAINTENANCE sheet (note: pick_zones returns 'Maintenance' key but rows are from MAINTENANCE sheet)
    if 'Maintenance' in assignments:
        emoji = FLOOR_EMOJI.get('Maintenance', '🏠')
        for row, task_name in assignments['Maintenance'].get('zones', []):
            mid = MID_BY_ROW.get(row)
            if mid is None:
                continue
            items.append({
                "source": "MAINTENANCE",
                "mid": mid,
                "row": row,            # reference only — NOT the key
                "label": f"{emoji} {task_name}",
                "emoji": emoji,
                "taskName": task_name,
                "uid": f"MAINTENANCE:{mid}",
            })

    # Business + Frog — these are TASKS rows surfaced via ZONE_PICKER's special selectors
    wb = load_workbook(beast_path, data_only=True)
    ws_t = wb["TASKS"]
    for special in ('Business', 'Frog'):
        if special not in assignments:
            continue
        emoji = FLOOR_EMOJI.get(special, '⭐')
        for row, label in assignments[special].get('zones', []):
            task_id = ws_t.cell(row, 1).value
            items.append({
                "source": "TASKS",
                "id": task_id,
                "row": row,
                "label": f"{emoji} {label}",
                "emoji": emoji,           # ← separate field for wheel UI
                "aw": assignments[special].get('aw'),
                "specialZone": special,
                "uid": f"TASKS:{task_id}",
            })
            claimed_rows.add(row)  # exclude from Daily Ten/Courage

    return items, claimed_rows


def _build_spin_wheel(beast_path):
    """All rows from the SPIN WHEEL sheet. ID-keyed via SID column."""
    wb = load_workbook(beast_path, data_only=True)
    ws = wb["SPIN WHEEL"]
    # Locate SID column by header (appended, schema-safe)
    H = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    sid_col = H.get("SID")
    items = []
    for r in range(2, ws.max_row + 1):
        label = ws.cell(r, 1).value
        if not label:
            continue
        sid = ws.cell(r, sid_col).value if sid_col else None
        if sid is None:
            # No SID — skip rather than emit an unsafe row-based uid
            continue
        items.append({
            "source": "SPIN_WHEEL",
            "sid": sid,
            "row": r,            # kept for reference/debug only — NOT used as key
            "label": label,
            "uid": f"SPIN:{sid}",
        })
    return items


# ─── PUBLIC API ──────────────────────────────────────────────────
def build_seed(beast_path=None, today=None, courage_micros=None):
    """Build the full 4-bucket seed for the chaos lotus wheel.

    Args:
      beast_path: path to local beast .xlsx (default: /home/claude/masterHiveBrain.xlsx)
      today: date object (default: today UTC — caller may want Alaska date)
      courage_micros: optional dict {parent_task_id: "🔥 first-step text"} to override
                     auto-generated micros for AW 1-2 tasks. If a courage pick's
                     parent ID isn't in this dict, a generic fallback is used.

    Returns: list of task dicts ready to push to state.json
    """
    beast_path = beast_path or DEFAULT_BEAST_PATH
    today = today or date.today()
    courage_micros = courage_micros or {}

    all_tasks = []

    # 1. Build zones first so we know which TASKS rows are claimed by Frog/Business
    zone_items, claimed_rows = _build_zones(beast_path, today)

    # 2. Load available tasks once for Daily Ten + Courage builds
    available = _load_available_tasks(beast_path, today)

    # 3. Daily Ten (excluding Frog/Business rows)
    daily_ten = _build_daily_ten(available, claimed_rows, today, beast_path=beast_path)
    for t in daily_ten:
        # Prepend ✨ to recurring rhythm tasks so they read distinctly on the wheel.
        _lbl = t["label"]
        if t.get("recurring") and not str(_lbl).startswith("✨"):
            _lbl = f"✨ {_lbl}"
        all_tasks.append({
            "source": "TASKS",
            "id": t["id"],
            "row": t["row"],
            "label": _lbl,
            "aw": t["aw"],
            "pri": t["pri"],
            "dur": t["dur"],
            "ml": t["ml"],
            "proj": t["proj"],
            "cat": t["cat"],
            "critical": t["critical"],
            "uid": f"TASKS:{t['id']}",
        })

    # 4. Courage (excluding Frog/Business rows)
    courage_picks = _build_courage(available, claimed_rows, today, courage_micros, beast_path=beast_path)
    for t in courage_picks:
        # NEW MODEL: no pre-drafted micro. 🔥 flame = decompose-cue. The label
        # is just the task; tapping it done means "I faced it." (courage_micros
        # kept as optional override for back-compat but unused by default.)
        micro = courage_micros.get(t["id"]) or courage_micros.get(t["row"])
        if micro:
            label = micro if micro.startswith("🔥") else f"🔥 {micro}"
        else:
            label = f"🔥 {t['label']}"
        all_tasks.append({
            "source": "COURAGE",
            "parentId": t["id"],
            "parentRow": t["row"],
            "parentLabel": t["label"],
            "stepIndex": 0,
            "aw": t["aw"],
            "due": t["due"].isoformat() if t["due"] else None,
            "label": label,
            "emoji": "🔥",
            "uid": f"COURAGE:{t['id']}:0",
        })

    # 5. Add zone items (computed in step 1)
    all_tasks.extend(zone_items)

    # 6. Spin Wheel items
    all_tasks.extend(_build_spin_wheel(beast_path))

    return all_tasks


def push_seed(tasks, commit_msg=None):
    """Push tasks list to state.json on github via chaos_kv_helper.

    Bumps version automatically. Returns the github commit result.
    """
    import chaos_kv_helper as ckv

    current = ckv.load_state()
    new_version = current.get("version", 0) + 1

    new_state = {
        "tasks": tasks,
        "version": new_version,
        "updated": datetime.now(timezone.utc).isoformat(),
        "buckets": _count_by_source(tasks),
    }

    msg = commit_msg or f"Seed wheel: {len(tasks)} items"
    return ckv.save_state(new_state, commit_msg=msg)


def seed_wheel(beast_path=None, today=None, courage_micros=None, commit_msg=None):
    """Convenience: build + push in one call."""
    tasks = build_seed(beast_path, today, courage_micros)
    result = push_seed(tasks, commit_msg)
    print(f"🪷 Seeded wheel: {len(tasks)} items")
    counts = _count_by_source(tasks)
    for src, ct in sorted(counts.items()):
        print(f"   {src}: {ct}")
    print(f"   commit: {result['commit']['sha'][:12]}")
    return tasks


def _count_by_source(tasks):
    counts = {}
    for t in tasks:
        src = t.get("source", "?")
        counts[src] = counts.get(src, 0) + 1
    return counts
